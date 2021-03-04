'xls writer'

from datetime import date
from decimal import Decimal
from typing import Dict
from typing import Iterable
from typing import List
from typing import Tuple
from typing import Union

from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pdf2xls.model import AdditionalDetail

from ..model import ColumnHeader
from ..model import Info
from .abcwriter import ABCWriter

E = Union[None, str, date, Decimal]

NUMBER_FORMAT_TEXT = '@'
NUMBER_FORMAT_DATE = 'mm-dd-yy'
NUMBER_FORMAT_NUMBER = '0.00'
NUMBER_FORMAT_CURRENCY = r'#,##0.00\ "€"'

VALUE_NUMBER_FORMAT = {
    ColumnHeader.periodo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.livello_categoria: NUMBER_FORMAT_NUMBER,
    ColumnHeader.n_scatti: NUMBER_FORMAT_NUMBER,
    ColumnHeader.minimo: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.scatti: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.superm: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.sup_ass: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.edr: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.totale_retributivo: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.ferie_a_prec: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_spett: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_godute: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_saldo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_a_prec: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_spett: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_godute: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_saldo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.netto_da_pagare: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.legenda_ordinario: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_straordinario: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_ferie: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_reperibilita: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_rol: NUMBER_FORMAT_NUMBER,
    ColumnHeader.detail: NUMBER_FORMAT_TEXT
}


class XlsWriter(ABCWriter):
    'write infos on an .xls'

    def write_infos(self, infos: List[Info]) -> None:
        'atomically write all the infos'

        # collect all details of all infos:
        details = {additional_detail.descrizione
                   for info in infos
                   for additional_detail in info.additional_details}

        # there are 1 + len(keys)-1 + len(set(details)) columns

        header: List[Tuple[E, str]] = [('month', NUMBER_FORMAT_TEXT)]
        for key in ColumnHeader:
            if key is not ColumnHeader.detail:
                header.append((key.name, NUMBER_FORMAT_TEXT))
            else:
                for detail in details:
                    header.append((detail, NUMBER_FORMAT_TEXT))

        rows: List[List[Tuple[E, str]]] = [header]

        order = {e: i for i, e in enumerate(ColumnHeader)}
        for info in infos:
            row: List[Tuple[E, str]] = [(info.when, NUMBER_FORMAT_DATE)]
            # sort columns by Key "order"
            for column in sorted(info.columns, key=lambda c: order[c.header]):
                row.append((column.howmuch,
                            VALUE_NUMBER_FORMAT[column.header]))

            # group additional_details by descrizione
            additional_details: Dict[str, AdditionalDetail] = {}
            for ad in info.additional_details:
                additional_details[ad.descrizione] = ad

            for detail in details:
                maybe_ad = additional_details.get(detail)
                # cosa esportare nell'xls?
                row.append((None, NUMBER_FORMAT_TEXT)
                           if maybe_ad is None
                           else (maybe_ad.competenze, NUMBER_FORMAT_NUMBER))
            rows.append(row)

        widths: Dict[str, int] = {}
        for row in rows:
            for i, cell in enumerate(row, start=1):
                column_letter = get_column_letter(i)
                widths[column_letter] = max(widths.get(column_letter, 0),
                                            2 * len(str(cell[0])))

        print(f'output of {len(rows)}x{len(rows[0])} cells')

        workbook = Workbook(write_only=True)
        try:
            # create the main sheet
            sheet = workbook.create_sheet('main')
            # first set the width of the columns
            for column_letter, width in widths.items():
                sheet.column_dimensions[column_letter].width = width
            # then add the data
            for row in rows:
                sheet.append([self._cell(sheet, *cell) for cell in row])
        finally:
            workbook.save(self.name)

    def _row(self, e: E, fmt: str, it: Iterable[Tuple[E, str]]) -> List[Tuple[E, str]]:
        row = []
        row.append((e, fmt))
        row.extend(it)
        return row

    def _cell(self, sheet: Worksheet, value: E, number_format: str) -> Cell:
        cell = Cell(sheet, value=value)
        cell.number_format = number_format
        return cell
