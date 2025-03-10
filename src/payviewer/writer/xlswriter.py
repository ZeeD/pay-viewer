from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING
from typing import Final

from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook.workbook import Workbook
from payviewer.model import ColumnHeader
from payviewer.model import Info
from payviewer.writer.abcwriter import ABCWriter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

E = None | str | date | Decimal

NUMBER_FORMAT_TEXT: Final = '@'
NUMBER_FORMAT_DATE: Final = 'mm-dd-yy'
NUMBER_FORMAT_NUMBER: Final = '0.00'
NUMBER_FORMAT_CURRENCY: Final = r'#,##0.00\ "€"'

VALUE_NUMBER_FORMAT = {
    ColumnHeader.periodo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.livello_categoria: NUMBER_FORMAT_NUMBER,
    ColumnHeader.n_scatti: NUMBER_FORMAT_NUMBER,
    ColumnHeader.minimo: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.scatti: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.superm: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.sup_ass: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.edr: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.totale_retributivo: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.ferie_a_prec: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_spett: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_godute: NUMBER_FORMAT_NUMBER,
    ColumnHeader.ferie_saldo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_a_prec: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_spett: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_godute: NUMBER_FORMAT_NUMBER,
    ColumnHeader.par_saldo: NUMBER_FORMAT_NUMBER,
    ColumnHeader.netto_da_pagare: NUMBER_FORMAT_CURRENCY,
    ColumnHeader.legenda_ordinario: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_straordinario: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_ferie: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_reperibilita: NUMBER_FORMAT_NUMBER,
    ColumnHeader.legenda_rol: NUMBER_FORMAT_NUMBER,
    ColumnHeader.detail: NUMBER_FORMAT_TEXT,
}


class XlsWriter(ABCWriter):
    def write_infos(self, infos: list[Info]) -> None:  # noqa: C901
        def clean(descrizione: str) -> str:
            for prefix in (
                'AD.COM.LE DA TR. NEL ',
                'AD.REG.LE DA TR. NEL ',
                'TICKET PASTO',
            ):
                if descrizione.startswith(prefix):
                    return f'{prefix}*'

            return descrizione

        # collect all details of all infos:
        details = sorted(
            {
                clean(additional_detail.descrizione)
                for info in infos
                for additional_detail in info.additional_details
            }
        )

        # there are 1 + len(keys)-1 + len(details) columns

        header: list[tuple[E, str]] = [('month', NUMBER_FORMAT_TEXT)]
        for column_header in ColumnHeader:
            if column_header is not ColumnHeader.detail:
                header.append((column_header.name, NUMBER_FORMAT_TEXT))
            else:
                header.extend(
                    (detail, NUMBER_FORMAT_TEXT) for detail in details
                )

        rows: list[list[tuple[E, str]]] = [header]

        for info in infos:
            # group columns by column_header
            columns = {column.header: column for column in info.columns}
            # group additional_details by descrizione
            additional_details = {
                clean(additional_detail.descrizione): additional_detail
                for additional_detail in info.additional_details
            }

            row: list[tuple[E, str]] = [(info.when, NUMBER_FORMAT_DATE)]
            for column_header in ColumnHeader:
                if column_header is not ColumnHeader.detail:
                    row.append(
                        (
                            columns[column_header].howmuch,
                            VALUE_NUMBER_FORMAT[column_header],
                        )
                        if column_header in columns
                        else (None, NUMBER_FORMAT_TEXT)
                    )
                else:
                    # cosa esportare nell'xls?
                    row.extend(
                        (
                            additional_details[detail].competenze
                            - additional_details[detail].trattenute,
                            NUMBER_FORMAT_NUMBER,
                        )
                        if detail in additional_details
                        else (None, NUMBER_FORMAT_TEXT)
                        for detail in details
                    )
            rows.append(row)

        widths: dict[str, int] = {}
        for row in rows:
            for i, cell in enumerate(row, start=1):
                column_letter = get_column_letter(i)
                widths[column_letter] = max(
                    widths.get(column_letter, 0), 2 * len(str(cell[0]))
                )

        workbook = Workbook(write_only=True)
        try:
            # create the main sheet
            sheet = workbook.create_sheet('main')
            # first set the width of the columns
            for column_letter, width in widths.items():
                sheet.column_dimensions[column_letter].width = width
            # then add the data
            for row in rows:
                sheet.append(
                    [
                        self._cell(sheet, value, number_format)
                        for value, number_format in row
                    ]
                )
        finally:
            workbook.save(self.name)

    def _cell(self, sheet: 'Worksheet', value: E, number_format: str) -> Cell:
        cell = Cell(sheet, value=value, column=1)
        cell.number_format = number_format
        return cell
